#!/usr/bin/env python3
"""
parse-price.py — Линкер прайс Excel → lib/products-data-new.ts
Использование: python3 scripts/parse-price.py path/to/price.xlsx
"""
import openpyxl, json, re, sys
from pathlib import Path

MARKUP = 1.45

BRAND_SHEETS = {
    "Церсанит 01.01.2026 v2":        {"brand": "Cersanit",         "slug_prefix": "cersanit",      "parser": "cersanit"},
    "Керама Марацци 12.01.2026":      {"brand": "Kerama Marazzi",   "slug_prefix": "kerama-marazzi","parser": "kerama"},
    "Азори 15.08.2025 v4":           {"brand": "Азори",            "slug_prefix": "azori",         "parser": "azori"},
    "Нефрит-Керамика 19.01.2026":    {"brand": "Нефрит-Керамика",  "slug_prefix": "nefrit",        "parser": "azori"},
    "Урал.гранит-Гранитея 01.01":    {"brand": "Урал Гранит",      "slug_prefix": "ural-granit",   "parser": "ural"},
    "Бонапарт 12.01.2026":           {"brand": "Бонапарт",         "slug_prefix": "bonaparte",     "parser": "napoleon"},
    "Грация Kерамика Шахты 01.01":   {"brand": "Грация Керамика",  "slug_prefix": "gracia",        "parser": "azori"},
    "Идальго 01.01.2026":            {"brand": "Идальго",          "slug_prefix": "idalgo",        "parser": "azori"},
    "Элетто 15.04.2024 v2":          {"brand": "Элетто",           "slug_prefix": "eletto",        "parser": "azori"},
    "М-Квадрат 13.01.2025":          {"brand": "М-Квадрат",        "slug_prefix": "m-kvadrat",     "parser": "azori"},
}

TRANSLIT_MAP = {
    "а":"a","б":"b","в":"v","г":"g","д":"d","е":"e","ё":"yo","ж":"zh","з":"z",
    "и":"i","й":"y","к":"k","л":"l","м":"m","н":"n","о":"o","п":"p","р":"r",
    "с":"s","т":"t","у":"u","ф":"f","х":"kh","ц":"ts","ч":"ch","ш":"sh","щ":"sch",
    "ъ":"","ы":"y","ь":"","э":"e","ю":"yu","я":"ya",
}
TRANSLIT_MAP.update({k.upper(): v.upper() for k,v in TRANSLIT_MAP.items()})

def slugify(text):
    t = "".join(TRANSLIT_MAP.get(c,c) for c in str(text or "")).lower().strip()
    t = re.sub(r"[^\w\s-]","",t); t = re.sub(r"[\s_]+","-",t)
    return re.sub(r"-+","-",t).strip("-")[:80]

def get_format(size):
    parts = str(size or "").replace(",",".").replace(" ","").split("x")
    if len(parts) >= 2: return f"{parts[0]}x{parts[1]}"
    parts2 = str(size or "").replace(" ","").split("*")
    if len(parts2) >= 2: return f"{parts2[0]}x{parts2[1]}"
    return str(size or "")

def detect_type(name):
    n = str(name).lower()
    if "мозаик" in n: return "Мозаика"
    if "ступен" in n: return "Ступени"
    if any(x in n for x in ["керамогранит","pgr","gres","грес"]): return "Керамогранит"
    return "Керамическая плитка"

def detect_color(name):
    for c in ["Белый","Черный","Серый","Бежевый","Коричневый","Синий","Голубой",
              "Зеленый","Красный","Розовый","Антрацит","Графит","Кремовый","Терракот","Мокко"]:
        if c.lower() in str(name).lower(): return c
    return "Микс"

def safe_float(v):
    try: return float(v) if v else 0
    except: return 0

def calc_price(price_in, price_retail):
    pr = safe_float(price_retail)
    pi = safe_float(price_in)
    if pr > 0: return round(pr * 0.8)
    if pi > 0: return round(pi * MARKUP)
    return 0

def make_p(sku, name, collection, size, price_in, price_retail, brand_info):
    price = calc_price(price_in, price_retail)
    if not name or not sku or price == 0: return None
    pr = int(safe_float(price_retail)) if safe_float(price_retail) > 0 else None
    pfx = brand_info["slug_prefix"]
    return {
        "id": f'{pfx}-{slugify(sku)}',
        "sku": str(sku).strip(),
        "name": str(name).strip(),
        "slug": f'{pfx}-{slugify(collection or "kollekciya")}-{slugify(sku)}'[:100],
        "brand": brand_info["brand"],
        "collection": str(collection or "").strip(),
        "product_type": detect_type(name),
        "format": get_format(size),
        "color": detect_color(name),
        "price_retail": price,
        "price_official": pr,
    }

def find_header(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(str(c).strip() in ("Артикул","Коллекция","Наименование") for c in row if c):
            return i
    return 5

def parse_cersanit(ws, bi):
    results, skip = [], find_header(ws)
    current_coll = ""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= skip: continue
        # Collection header rows have text in col 1 but no price
        if row[1] and not str(row[1]).strip().isdigit() and not row[10]:
            current_coll = str(row[1]).strip() if row[1] else current_coll
            continue
        sku = str(row[1]).strip() if row[1] else ""
        if not sku.isdigit(): continue
        coll = str(row[5]).strip() if row[5] else current_coll
        p = make_p(sku, row[7], coll, row[8], row[10], row[11], bi)
        if p: results.append(p)
    return results

def parse_kerama(ws, bi):
    results, skip = [], find_header(ws)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= skip: continue
        sku = str(row[1]).strip() if row[1] else ""
        if not sku.isdigit(): continue
        discont = str(row[15]).strip() if len(row) > 15 and row[15] else ""
        if discont.lower() in ("да","yes","снят"): continue
        p = make_p(sku, row[5], row[4], row[6], row[10], row[11], bi)
        if p: results.append(p)
    return results

def parse_azori(ws, bi):
    # cols: артикул(1), тип(3), наименование(4), размер(5), вход_коробками(8), розница(9)
    results, skip = [], find_header(ws)
    current_coll = ""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= skip: continue
        sku_raw = row[1]
        if not sku_raw: continue
        sku = str(sku_raw).strip()
        # Collection header: text in col 1, no name
        if sku and not row[4]:
            current_coll = re.sub(r" коллекция.*","",sku,flags=re.I).strip()
            continue
        # Skip format headers like "Формат 31,5х63"
        if "формат" in sku.lower(): continue
        name = row[4] if row[4] else ""
        if not name: continue
        p = make_p(sku, name, current_coll, row[5], row[8], row[9], bi)
        if p: results.append(p)
    return results

def parse_ural(ws, bi):
    results, skip = [], find_header(ws)
    current_coll = ""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= skip: continue
        if not row[1]: continue
        sku = str(row[1]).strip()
        if not sku.isdigit():
            if not row[5]: current_coll = sku
            continue
        p = make_p(sku, row[4], current_coll or row[3], row[3], row[6], row[7], bi)
        if p: results.append(p)
    return results

def parse_napoleon(ws, bi):
    # Бонапарт: артикул(2), размер(3/4), наименование(2 is name), вход(6), розница(7)
    results, skip = [], find_header(ws)
    current_coll = ""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= skip: continue
        if not row[2]: continue
        art = str(row[2]).strip()
        if not row[6]:  # no price = section header
            current_coll = art
            continue
        size = str(row[4]).strip() if row[4] else str(row[3]).strip() if row[3] else ""
        p = make_p(art, art, current_coll, size, row[6], row[7], bi)
        if p: results.append(p)
    return results

PARSERS = {
    "cersanit": parse_cersanit,
    "kerama": parse_kerama,
    "azori": parse_azori,
    "ural": parse_ural,
    "napoleon": parse_napoleon,
}

def main():
    price_path = sys.argv[1] if len(sys.argv) > 1 else None
    if not price_path:
        for c in [Path(__file__).parent/"Копия_Праи_сы_-_вход.xlsx", Path.home()/"Downloads"/"Копия_Праи_сы_-_вход.xlsx"]:
            if c.exists(): price_path = str(c); break
    if not price_path or not Path(price_path).exists():
        print("Использование: python3 scripts/parse-price.py path/to/price.xlsx"); sys.exit(1)

    print(f"📂 Читаю: {Path(price_path).name}")
    wb = openpyxl.load_workbook(price_path, read_only=True, data_only=True)

    all_products, stats = [], {}
    for sheet_name, bi in BRAND_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Лист не найден: {sheet_name}"); continue
        parser_fn = PARSERS[bi["parser"]]
        products = parser_fn(wb[sheet_name], bi)
        stats[bi["brand"]] = len(products)
        all_products.extend(products)
        print(f"  ✅ {bi['brand']}: {len(products)} товаров")

    seen, unique = set(), []
    for p in all_products:
        s = p["slug"]
        if s in seen: p["slug"] = s + "-" + p.get("sku","")[-4:]
        seen.add(p["slug"]); unique.append(p)

    print(f"\n📊 Итого: {len(unique)} товаров из {len([v for v in stats.values() if v>0])} брендов")

    def tv(v):
        if v is None: return "undefined"
        if isinstance(v,bool): return "true" if v else "false"
        if isinstance(v,(int,float)): return str(v)
        return json.dumps(v,ensure_ascii=False)

    def pts(p):
        fields = ["id","sku","name","slug","brand","collection","product_type","format","color","price_retail","price_official"]
        return "{\n" + "\n".join(f"  {f}: {tv(p.get(f))}," for f in fields if p.get(f) is not None) + "\n}"

    out = Path(__file__).parent.parent/"lib"/"products-data-new.ts"
    out.write_text(
        f"""// AUTO-GENERATED by scripts/parse-price.py — НЕ РЕДАКТИРОВАТЬ ВРУЧНУЮ
// Товаров: {len(unique)} | {json.dumps(stats, ensure_ascii=False)}

export const importedProducts = [
{",\n".join(pts(p) for p in unique)}
] as const
""", encoding="utf-8")
    print(f"✅ {out}")
    print("\nДалее:")
    print("  1. Проверь lib/products-data-new.ts → переименуй в products-data-imported.ts")
    print("  2. В products-data.ts: import { importedProducts } from ./products-data-imported")
    print("  3. export const products = [...localProducts, ...importedProducts]")

if __name__ == "__main__":
    main()
